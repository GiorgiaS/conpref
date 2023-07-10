# This class checks the edge.xlsx file to find the edge node
# serving the user according to the user's coordinates.
# It assigns a random coverage range to each edge equals to 500m.
# Moreover, it extracts the closer edge to the device

from openpyxl import load_workbook

from pygeodesy.sphericalNvector import LatLon


class EdgeLocator:
    def getEdgeId(self, coordinates):
        edgeID = 0
        currentDistance = 0

        userPosition = LatLon(coordinates[0], coordinates[1])

        file = load_workbook('./melbourne_telcos.xlsx')

        sheet = file['testing_area']
        sheet = sheet[2: sheet.max_row]

        for val in sheet:
            edgePoint = LatLon(val[1].value, val[2].value)
            # print("EdgeLocator::getEdgeId - distance: ", edgePoint.distanceTo(userPosition))
            distance = edgePoint.distanceTo(userPosition)
            if(distance < 500 and (currentDistance > distance or currentDistance == 0)):
                currentDistance = distance
                edgeID = val[0].value
                # print("EgeLocator::getEdgeId - distance: ", distance)
                # print("EdgeLocator::getEdgeId - coordinates: ", edgeID, val[1].value, val[2].value)
        return edgeID
